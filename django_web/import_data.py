from lectures.models import Lecture
import openpyxl

wb = openpyxl.load_workbook('data/all_lecture.xlsx')
ws = wb.active

for row in range(2,ws.max_row+1):
    department = ws["A{}".format(row)].value
    number = ws["B{}".format(row)].value
    name = ws["C{}".format(row)].value
    teacher = ws["D{}".format(row)].value
    time_and_room = ws["E{}".format(row)].value
    remark = ws["F{}".format(row)].value
    
    if row % 1000 == 0: print(row)
    Lecture.objects.create(department=department, number=number, name=name, teacher=teacher, time_and_room=time_and_room, remark=remark)
