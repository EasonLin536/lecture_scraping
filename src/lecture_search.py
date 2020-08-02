import sys
import openpyxl

# lecture file
wb = openpyxl.load_workbook(sys.argv[1])
ws = wb.active
# outfile = openpyxl.load_workbook(sys.argv[2])


# Clear outfile
# def clear_excel():
#     for sheet in outfile.sheetnames:
#         del outfile[sheet]


# Create outfile sheet
# def create_sheet():
#     sheet_name = "results"
#     outfile.create_sheet(sheet_name)
#     ws = outfile[sheet_name]
#     ws.append(["授課對象", "課名", "課名", "授課教師", "時間教室", "備註"])


# Process command
def command():
    cmd = input("search>> ")
    if cmd == "quit": return False
    elif cmd == "": 
        print("example: [-de 電機] [-nu EE1] [-na 電磁學一] [-te 江衍偉] [-ti 一6,7,8,9] [-ro 電二225] [-re 兼通識]")
        print("type \"quit\" to exit program")
        return
    tokens = cmd.split()

    department = ""
    number = ""
    name = ""
    teacher = ""
    time = ""
    room = ""
    remark = ""

    for i in range(0,len(tokens),2):
        if tokens[i] == "-de":
            department = tokens[i+1]
        elif tokens[i] == "-nu":
            number = tokens[i+1]
        elif tokens[i] == "-na":
            name = tokens[i+1]
        elif tokens[i] == "-te":
            teacher = tokens[i+1]
        elif tokens[i] == "-ti":
            time = tokens[i+1]
        elif tokens[i] == "-ro":
            room = tokens[i+1]
        elif tokens[i] == "-re":
            remark = tokens[i+1]
        else: 
            print("Illegal command.")
            return
    
    search(department, number, name, teacher, time, room, remark)


# Search keyword in certain column
def search_str(mode, target, lecture_list):
    new_list = []
    for row in lecture_list:
        column = mode # column of room and time
        cell_name = "{}{}".format(column, row)
        if ws[cell_name].value != None:
            if ws[cell_name].value.find(target) != -1:
                new_list.append(row)

    return new_list


# Print out info and append to ws in certain row
def print_class(row):
    department =  ws["A{}".format(row)].value
    number = ws["B{}".format(row)].value
    name = ws["C{}".format(row)].value
    teacher = ws["D{}".format(row)].value
    time_and_room = ws["E{}".format(row)].value
    remarks = ws["F{}".format(row)].value
        
    print("{} {} {} {} {} {}".format(department, number, name, teacher, time_and_room, remarks))
    # outsheet = outfile.active
    # outsheet.append([department, number, name, teacher, time_and_room, remarks])


# Functions for search time
# Turn 一 to 1 ,etc
def chin_to_num(chin_day):
    if chin_day == "一": return 1
    elif chin_day == "二": return 2
    elif chin_day == "三": return 3
    elif chin_day == "四": return 4
    elif chin_day == "五": return 5
    else: 
        print("Illegal time!")
        return False
    

# Turn 一6,7,8 into [6,7,8]
def extract_time(time_str, day=0):
    if len(time_str) == 1: return ['1','2','3','4','5','6','7','8','9','10','A','B','C','D']
    if day != 0: # Looking for cirtain day
        if day == 1: str = "一"
        elif day == 2: str = "二"
        elif day == 3: str = "三"
        elif day == 4: str = "四"
        else: str = "五"
    
        begin = time_str.find(str)
        end = time_str.find("(", begin)
        time_str = time_str[begin:end]
    
    return time_str[1:].split(',')


# Clear the room to avoid mistake
def process_str(time_str):
    begin = 0
    end = 0
    tmp = ""
    while True:
        begin = time_str.find("(", begin) + 1
        tmp += time_str[end:begin]
        end = time_str.find(")", end) + 1
        if end == 0: break
    
    return tmp
# End for funcs. of search time

# return a new list of rows that fits certain time
def search_time(time, lecture_list):
    day = chin_to_num(time[0])
    if day is False: main()

    new_list = []
    for row in lecture_list:
        column = 'E' # column of room and time
        cell_name = "{}{}".format(column, row)

        time_str = process_str(ws[cell_name].value)

        all_times = extract_time(time)
        # class_times = extract_time(ws[cell_name].value, day)
        class_times = extract_time(time_str, day)
                
        if set(class_times).issubset(set(all_times)):
            new_list.append(row)

    return new_list


# Handle every kind of search
def search(department, number, name, teacher, time, room, remark):
    lecture_list = range(1,ws.max_row+1)
    # remark
    if remark:
        print("Searching \"{}\" in remarks...".format(remark))
        lecture_list = search_str("F", remark, lecture_list)
    # time: 一6,7,8
    if time:
        print("Searching {} on 星期{}...".format(time[1:], time[0]))
        lecture_list = search_time(time, lecture_list)
    # name
    if name:
        print("Searching lecture: {}...".format(name))
        lecture_list = search_str("C", name, lecture_list)
    # teacher
    if teacher:
        print("Searching teacher: {}...".format(teacher))
        lecture_list = search_str("D", teacher, lecture_list)
    # number
    if number:
        print("Searching number: {}...".format(number))
        lecture_list = search_str("B", number, lecture_list)
    # room
    if room:
        print("Searching room: {}...".format(room))
        lecture_list = search_str("E", room, lecture_list)
    # department
    if department:
        print("Searching department: {}...".format(department))
        lecture_list = search_str("A", department, lecture_list)

    for row in lecture_list:
        print_class(row)
    
    print("{} result(s) found".format(len(lecture_list)))


def main():
    while True:
        # clear_excel()
        # create_sheet()
        if command() is False: sys.exit(0)


main()