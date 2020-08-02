import sys
import requests
import webbrowser
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
import time

# Global variables
page_cnt = 3000
sem = sys.argv[1]
url = "http://nol.ntu.edu.tw/nol/coursesearch/search_result.php"
file_name = sys.argv[2]

wb = Workbook()


def create_sheet(day):
    sheet_name = "lecture sheet"
    wb.create_sheet(sheet_name)
    ws = wb[sheet_name]
    ws.append(["授課對象", "課號", "課名", "授課教師", "時間教室", "備註"])



def get_url(day):
    print("Getting url_page...")
    day_str = "week" + str(day)
    my_param = {
        'current-sem': sem,
        'cstype': '1',
        'alltime': 'no',
        day_str: '1',
        'allproced': 'yes',
        'allsel': 'yes',
        'page_cnt': str(page_cnt),
    }
    url_page = requests.get(url, params=my_param)
    url_page.encoding ='big5' # big5 expression
    return url_page.text


def scraping(day):
    url_page = get_url(day)
    soup = BeautifulSoup(url_page, 'html.parser')
    num_of_class = soup.find_all("table")[5].find("td").font.text.strip()
    num_of_class = int(num_of_class)
    
    table = soup.find_all("table")[6]
    columns = table.find_all("tr")

    ws = wb["lecture sheet"]
    if num_of_class > page_cnt:
        num_of_class = page_cnt
    print("Scraping " + str(num_of_class) + " classes")
    
    for i in range(1,num_of_class):
        column = columns[i]
        grids = column.find_all("td")
        
        department    = grids[1].text.strip()
        class_num     = grids[2].text.strip()
        class_name    = grids[4].text.strip()
        class_teacher = grids[10].text.strip()
        time_and_room = grids[12].text.strip()
        remarks       = grids[15].text.strip()
        
        ws.append([department, class_num, class_name, class_teacher, time_and_room, remarks])
        

def main():
    create_sheet(0)
    for day in range(1,6):
        print("day:" + str(day))
        scraping(day)    
    wb.save(file_name)


main()